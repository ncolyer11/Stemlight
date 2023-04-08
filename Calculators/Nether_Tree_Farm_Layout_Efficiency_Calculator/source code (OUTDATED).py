# THIS IS OUTDATED AND PURELY JUST TO SHOW HOW COOL V3* IS COMPARED TO V1 (this file)
from litemapy import Schematic
import openpyxl
import time

# I'm making a dedicated folder in .minecraft/schematics and then either running this code from
# there or adding a path to the 'schematic_name' variable
# MAKE sure you're selecting from the smallest numbers in 'corner 1' to larger numbers in 'corner 2'
# in the future i'll remap the input corner values so you don't have to worry about this but for now
# it'll give an out of range error

# change the schematic name from 'layout' to whatever you've named you're schematic
schematic_name = "layout.litematic"
workbook = openpyxl.load_workbook('heatmaps.xlsx')
cell_values = {}

start_time = time.time()


for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    for row in sheet.iter_rows():
        for cell in row:
            cell_values[(sheet_name, cell.row, cell.column)] = cell.value


def get_cell_value(sheet_name, row_number, column_number):
    return cell_values[(sheet_name, row_number, column_number)]


def stems_and_shrooms(avg_stems, avg_shroomlights, g, f):
    avg_stems += get_cell_value('stems', g, f)
    avg_shroomlights += get_cell_value('shroomlights', g, f)
    print(f"\t- Stems = {get_cell_value('stems', g, f)}\n\t- Shroomlights = "
          f"{get_cell_value('shroomlights', g, f)}")
    return avg_stems, avg_shroomlights


avg_shroomlights = 0
avg_wart_blocks = 0
avg_stems = 0

schem = Schematic.load(f"{schematic_name}")
reg = list(schem.regions.values())[0]

print(f'In: {reg.xrange()} {reg.yrange()} {reg.zrange()}\n')
for y in reg.yrange():
    for x in reg.xrange():
        for z in reg.zrange():
            b = reg.getblock(x, y, z)

            if b.blockid != "minecraft:air":
                f = x + 1 + (7 * z)
                g = 27 - y
                print(f"{x} {y} {z}:")
                if b.blockid == "minecraft:red_concrete":
                    avg_stems, avg_shroomlights = stems_and_shrooms(avg_stems, avg_shroomlights, g, f)
                    avg_wart_blocks += get_cell_value('vrm0', g, f)
                    print(f"\t- VRM0 = {get_cell_value('vrm0', g, f)}")
                elif b.blockid == "minecraft:blue_concrete":
                    avg_stems, avg_shroomlights = stems_and_shrooms(avg_stems, avg_shroomlights, g, f)
                    avg_wart_blocks += get_cell_value('vrm1', g, f)
                    print(f"\t- VRM1 = {get_cell_value('vrm1', g, f)}")
                elif b.blockid == "minecraft:cyan_concrete":
                    avg_stems, avg_shroomlights = stems_and_shrooms(avg_stems, avg_shroomlights, g, f)
                    avg_wart_blocks += get_cell_value('vrm2', g, f)
                    print(f"\t- VRM2 = {get_cell_value('vrm2', g, f)}")
                    avg_stems, avg_shroomlights = stems_and_shrooms(avg_stems, avg_shroomlights, g, f)
                elif b.blockid == "minecraft:light_blue_concrete":
                    avg_stems, avg_shroomlights = stems_and_shrooms(avg_stems, avg_shroomlights, g, f)
                    avg_wart_blocks += get_cell_value('vrm3', g, f)
                    print(f"\t- VRM3 = {get_cell_value('vrm3', g, f)}")
                else:
                    print(f"\t- INVALID blockstate ({b.blockid})")

print(f'\nAvg blocks harvested per cycle:\n\t- Stems: {avg_stems}\n\t- Shroomlights: '
      f'{avg_shroomlights}\n\t- Wart Blocks: {avg_wart_blocks}')

end_time = time.time()
elapsed_time = end_time - start_time
print("\nElapsed time:", elapsed_time, "seconds")
