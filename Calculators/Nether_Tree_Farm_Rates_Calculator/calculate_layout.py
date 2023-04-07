"""
program that takes an input of a 7x27x7 (XYZ) schematic in the .litematic format, representing the blocks a given
nether tree farm layout harvests per cycle, and decodes the 3D data inside that file to compare it to pre-calculated
heatmaps in a .xlsx file to work out the average stems, shroomlights and wart blocks harvested per cycle for that
nether tree farm layout
"""

# if the below import functions are underlined in red run `pip install litemapy` and `pip install openpyxl`
from litemapy import Schematic
import openpyxl


# taking file input from user and checking to see if the path and schematic size is valid
def schematic_to_efficiency(path, hat_cycles, trunk_cycles):
    schem = Schematic.load(f"{path}")
    reg = list(schem.regions.values())[0]

    workbook = openpyxl.load_workbook('heatmaps.xlsx')
    cell_values = {}

    # opening the heatmap excel spreadsheet once and writing all the values to an array
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                cell_values[(sheet_name, cell.row, cell.column)] = cell.value

    # function for accessing a single value/cell in the above array
    def get_cell_value(sheet_name2, row_number, column_number):
        return cell_values[(sheet_name2, row_number, column_number)]

    # function that computes the instantaneous value of the avg stems and shroomlights of the layout
    def stems_and_shrooms(average_stems, average_shroomlights, row2, col2, hat_cycles2, trunk_cycles2):
        average_stems += comp_probability(get_cell_value('stems', row2, col2), trunk_cycles2)
        average_shroomlights += comp_probability(get_cell_value('shroomlights', row2, col2), hat_cycles2)
        return average_stems, average_shroomlights

    def comp_probability(success_chance, attempts):
        return 1 - (1 - success_chance) ** attempts

    # initialising variables
    Xrange, Yrange, Zrange = reg.xrange(), reg.yrange(), reg.zrange()
    avg_shroomlights, avg_wart_blocks, avg_stems = 0, 0, 0

    for y in Yrange:
        for x in Xrange:
            for z in Zrange:
                block = reg.getblock(x, y, z)
                X, Y, Z = x, y, z
                if min(Xrange) < 0:
                    X += 6
                if min(Yrange) < 0:
                    Y += 26
                if min(Zrange) < 0:
                    Z += 6

                # 3D data is stored in 'slices' on a 2D spreadsheet, hence some math is needed to convert between them
                col = X + (7 * Z) + 1
                row = 27 - Y

                # setting vrm value
                if block.blockid == "minecraft:air":
                    continue
                elif block.blockid == "minecraft:red_concrete":
                    vrm = 'vrm0'
                elif block.blockid == "minecraft:blue_concrete":
                    vrm = 'vrm1'
                elif block.blockid == "minecraft:cyan_concrete":
                    vrm = 'vrm2'
                elif block.blockid == "minecraft:light_blue_concrete":
                    vrm = 'vrm3'

                # calculating stems and shroomlight values
                avg_stems, avg_shroomlights = stems_and_shrooms(avg_stems, avg_shroomlights, row, col,
                                                                hat_cycles, trunk_cycles)
                # calculating wart block values
                if X == 0 and Z == 0:
                    avg_wart_blocks += comp_probability(get_cell_value(vrm, row, col), trunk_cycles)
                else:
                    avg_wart_blocks += comp_probability(get_cell_value(vrm, row, col), hat_cycles)

    # finding the efficiencies of the layout factoring in the varying cycles
    stem_E = trunk_cycles * avg_stems * 24 / 221
    shroomlight_E = hat_cycles * avg_shroomlights / 2.03192455026454
    wart_block_E = hat_cycles * avg_wart_blocks / 62.045721996296 + trunk_cycles * avg_wart_blocks / 0.97951

    return avg_stems, avg_shroomlights, avg_wart_blocks, stem_E, shroomlight_E, wart_block_E
